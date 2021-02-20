# -*- coding: utf-8 -*-

from __future__ import print_function
import pickle
import io
import shutil
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient import errors
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

import os
import time
from datetime import datetime, date, timedelta
import calendar
from uuid import uuid4
import requests
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import csv
import win32com.client
import pandas as pd
import json

from PIL import Image, ImageDraw, ImageFont
import itertools

import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

# Source for getting edit form response url:
# https://xfanatical.com/blog/how-to-edit-google-forms-responses-in-the-spreadsheet/#source-code

# If modifying these scopes, delete the file token.pickle.
#SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']  # Read only
SHEET_SCOPES = ['https://www.googleapis.com/auth/spreadsheets']  # Read, edit, create and delete
DRIVE_SCOPES = ['https://www.googleapis.com/auth/drive']  # Read, edit, create and delete

# The ID and range of a sample spreadsheet.
#SPREADSHEET_ID = '1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms'  # Example
#RANGE_NAME = 'Class Data!A2:E'  # Example
SPREADSHEET_ID = <SPREADSHEET_ID>
SHEET_ID = <SHEET_ID>
SPREADSHEET_CORRECT_ID = <SPREADSHEET_CORRECT_ID>
RANGE_NAME = 'Form responses 1'

# Expense Claim Form Excel Template
TEMPLATE = <TEMPLATE_FILENAME>
PDF_PRINT_AREA = 'A1:F28'


def createPath(path):
    # Create path if not exists
    if not os.path.exists(path):
        os.mkdir(path)


def getCredential(config_path, api_type):
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists(os.path.join(config_path, api_type+'_token.pickle')):
        with open(os.path.join(config_path, api_type+'_token.pickle'), 'rb') as token:
            creds = pickle.load(token)
    
    # If there are no (valid) credentials available, let the user log in.
    if api_type == 'sheet':
        SCOPES = SHEET_SCOPES
    elif api_type == 'drive':
        SCOPES = DRIVE_SCOPES
    
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                os.path.join(config_path, api_type+'_credentials.json'), SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open(os.path.join(config_path, api_type+'_token.pickle'), 'wb') as token:
            pickle.dump(creds, token)
    return creds


def getSheetData(service, correct=False):
    if correct:
        result = service.spreadsheets().values().get(spreadsheetId=SPREADSHEET_CORRECT_ID,
                                     range=RANGE_NAME).execute()
    else:
        result = service.spreadsheets().values().get(spreadsheetId=SPREADSHEET_ID,
                                     range=RANGE_NAME).execute()
    values = result.get('values', [])
    
    response_list = []
    if not values:
        print('No data found.')
        return
    else:
        for index, row in enumerate(values):
            if row:
                if index == 0:
                    header_row = row
                else:
                    response_list.append(row)
            else:
                print('No row found.')
    return header_row, response_list


def downloadDrive(config_path, save_path, url_dict, claim_num, img_type):
    # Get Google Drive API credential
    creds = getCredential(config_path, 'drive')
    
    # Call the Drive API
    service = build('drive', 'v3', credentials=creds)
    
    for item_num, url in url_dict.items():
        file_id = url.split('=')[1]
        request = service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            try:
                status, done = downloader.next_chunk()
            except errors.HttpError as error:
                print('An error occurred while downloading from Google Drive: %s' % error)
                continue
            if not done:
                print('Cannot download file -> Claim Number:', claim_num, '| Item:', item_num)
                continue
                #print('Claim Number:', claim_num, '| Item:', item_num, '| Download %d%%.' % int(status.progress() * 100))
        
        # The file has been downloaded into RAM, now save it in a file
        fh.seek(0)
        with open(os.path.join(save_path, img_type+'_'+claim_num+'_'+item_num+'.jpg'), 'wb') as f:
            shutil.copyfileobj(fh, f)#, length=131072)


def deleteDrive(config_path, fileID):
    # Get Google Drive API credential
    creds = getCredential(config_path, 'drive')
    
    # Call the Drive API
    service = build('drive', 'v3', credentials=creds)
    
    # Delete file
    result = service.files().delete(fileId=fileID).execute()


def deleteFile(filepath, claim_num, file_type, url_dict=None):
    if file_type == 'Image':
        suffix = '.jpg'
        for filename in os.listdir(filepath):
            if filename.endswith(suffix) and filename[:-len(suffix)].split('_')[1] == claim_num:
                if url_dict:
                    if filename[:-len(suffix)].split('_')[2] not in url_dict:
                        os.remove(os.path.join(filepath, filename))
                else:
                    os.remove(os.path.join(filepath, filename))
    elif file_type == 'PDF':
        suffix = '.pdf'
        for filename in os.listdir(filepath):
            if filename.endswith(suffix) and filename[:-len(suffix)].split('_')[1] == claim_num:
                os.remove(os.path.join(filepath, filename))
    elif file_type == 'Excel':
        suffix = '.xlsx'
        for filename in os.listdir(filepath):
            if filename.endswith(suffix) and filename[:-len(suffix)].split('_')[1] == claim_num:
                os.remove(os.path.join(filepath, filename))


def updateCell(service, spreadsheet_id, range_name, value):
    values = [
        [
            value  # Cell values ...
        ],
        # Additional rows ...
    ]
    body = {
        'values': values
    }
    result = service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=range_name,
        valueInputOption='USER_ENTERED',
        body=body).execute()
    #print('{0} cells updated.'.format(result.get('updatedCells')))


def deleteCell(service, spreadsheet_id, range_name):
    clear_values_request_body = {
            # TODO: Add desired entries to the request body.
    }
    
    result = service.spreadsheets().values().clear(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            body=clear_values_request_body).execute()


def deleteRow(service, spreadsheet_id, row_num):
    spreadsheet_data = [
        {
            "deleteDimension": {
                "range": {
                    "sheetId": SHEET_ID,
                    "dimension": "ROWS",
                    "startIndex": row_num,
                    "endIndex": row_num+1
                }
            }
        }
    ]
    
    update_data = {"requests": spreadsheet_data}
    
    result = service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=update_data).execute()


def convertCurrency(currency, convert_date):
    url = 'https://api.exchangerate.host/convert?from='+currency+'&to=HKD&date='+convert_date
    try:
        response = requests.get(url)
        data = response.json()
    except requests.exceptions.RequestException as e:
        raise SystemExit(e)

    if data['success']:
        return data['result']


def excelToPDF(excel_path, pdf_path):
    app = win32com.client.Dispatch('Excel.Application')
    app.Visible = False
    wb = app.Workbooks.Open(excel_path)
    ws_index_list = [1] # Sheets to print
    for index in ws_index_list:
        # off-by-one so the user can start numbering the worksheets at 1
        ws = wb.Worksheets[index - 1]
        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesTall = 1
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.PrintArea = PDF_PRINT_AREA
        ws.Columns.AutoFit()
    
    wb.WorkSheets(ws_index_list).Select()
    wb.ActiveSheet.ExportAsFixedFormat(0, pdf_path)
    wb.Close(True)


def writeOverallSummary(summary_path, header_row, response):
    # Write overall summary
    overall_summary_directory = os.path.join(summary_path, 'summary_overall.csv')
    file_exists = os.path.isfile(overall_summary_directory)
    with open(overall_summary_directory, 'a', newline='\n') as csv_file:
        writer = csv.writer(csv_file)
        # Write header if overall summary does not exist
        if not file_exists:
            writer.writerow(header_row)
        writer.writerow(response)


def writeMonthSummary(summary_path, year, month, header_row, response):
    month_summary_directory = os.path.join(summary_path, 'summary_'+str(year)+'-'+str(month)+'.csv')
    file_exists = os.path.isfile(month_summary_directory)
    with open(month_summary_directory, 'a', newline='\n') as csv_file:
        writer = csv.writer(csv_file)
        # Write header if month summary does not exist
        if not file_exists:
            writer.writerow(header_row)
        writer.writerow(response)


def writeAuditSummary(summary_path, current_year, next_year, header_row, response):
    audit_summary_directory = os.path.join(summary_path, 'summary_audit_'+str(current_year)+'-'+str(next_year)+'.csv')
    file_exists = os.path.isfile(audit_summary_directory)
    with open(audit_summary_directory, 'a', newline='\n') as csv_file:
        writer = csv.writer(csv_file)
        # Write header if audit summary does not exist
        if not file_exists:
            writer.writerow(header_row)
        writer.writerow(response)


def writeClaimForm(service, index, config_path, template_path, form_path, response_dict, update=False):
    expense_date_header = ['Date of Expense-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
    description_header = ['Description-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
    currency_header = ['Currency-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
    expense_amount_header = ['Expense Amount-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
    HKD_amount_header = ['Expense Amount in HKD-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
    evidence_header = ['Evidence of Exchange Rate-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
    receipt_header = ['Upload Receipts-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
    other_item_header = ['Do you have other expense items to claim?-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
    delete_item_header = ['Do you wish to delete expense item?-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
    
    # Assign version number 1
    ver_num = 1
    
    if update:
        # Get previous version of Excel Claim Form
        for filename in os.listdir(form_path):
            if filename.endswith('.xlsx') and filename[:-5].split('_')[1] == response_dict['Claim Number']:
                book = openpyxl.load_workbook(os.path.join(form_path, filename), data_only=True)
                sheet = book.active
                ver_num = int(sheet.cell(row=1, column=5).value) + 1
                book.close()
    
    # Create new expense claim output
    book = openpyxl.load_workbook(os.path.join(template_path, TEMPLATE))
    sheet = book.active
    
    # Apply data validation for status
    # Create a data-validation object with list validation
    dv = DataValidation(type='list', formula1='"Approved,Rejected"', allow_blank=False)
    # Optionally set a custom error message
    dv.error = 'Please choose the status from the drop-down list.'
    dv.errorTitle = 'Invalid Entry'
    # Optionally set a custom prompt message
    dv.prompt = 'Please choose a status for this item.'
    dv.promptTitle = 'Status'
    # Apply the validation to a range of cells
    dv.add('F10:F22')
    # Add the data-validation object to the worksheet
    sheet.add_data_validation(dv)
    
    # Update version number
    sheet.cell(row=1, column=5).value = ver_num
    
    # Write contents
    receipt_dict = {}
    evidence_dict = {}
    has_row = False  # To record if the claim has any row
    delete_num = ''  # To record the item number to delete
    delete_list = []  # To reorder the items
    for key, value in response_dict.items():
        # Break out if it is the last item
        if key in other_item_header:
            if value == 'No. I am done with the expense details.':
                break
            continue
        # Write name
        if key == 'Name':
            # Write sheet title
            sheet.title = 'Form_'+ value
            sheet.cell(row=3, column=2).value = value
            continue
        # Write location
        if key == 'Location':
            sheet.cell(row=4, column=2).value = value
            continue
        # Write project code
        if key == 'Project Code':
            sheet.cell(row=5, column=2).value = value
            continue
        # Write submission date
        if key == 'Timestamp':
            submission_date = datetime.strptime(value, '%d/%m/%Y %H:%M:%S').strftime('%d/%m/%Y')
            sheet.cell(row=6, column=2).value = submission_date
            continue
        # Check if the item is deleted and record item number for later usage
        if key in delete_item_header:
            if value == 'Yes':
                delete_num = key[-2:]
                delete_list.append(delete_num)
        # Write expense date and item, store date for exchange rate
        if key in expense_date_header:
            if value and key[-2:] != delete_num:
                sheet.cell(row=9+int(key[-2:])-len(delete_list), column=1).value = str(int(key[-2:])-len(delete_list)).zfill(2)
                sheet.cell(row=9+int(key[-2:])-len(delete_list), column=2).value = value
                expense_date = datetime.strptime(value, '%d/%m/%Y').strftime('%Y-%m-%d')
                has_row = True
            continue
        # Write description
        if key in description_header:
            if value and key[-2:] != delete_num:
                sheet.cell(row=9+int(key[-2:])-len(delete_list), column=3).value = value
            continue
        # Write currency
        if key in currency_header:
            if value and key[-2:] != delete_num:
                # Write HKD anyways
                sheet.cell(row=9+int(key[-2:])-len(delete_list), column=4).value = 'HKD'
                currency = value
                # Record exchange rate of foreign currency
                if currency and currency != 'HKD':
                    exchange_rate = convertCurrency(currency, expense_date)
                    # Use submission date if user enter date in the future
                    if not exchange_rate:
                        exchange_rate = convertCurrency(currency, datetime.strptime(submission_date, '%d/%m/%Y').strftime('%Y-%m-%d'))
            continue
        # Write expense amount
        if key in expense_amount_header:
            if value and key[-2:] != delete_num:
                # Convert amount in foreign currency and record it
                if currency and currency != 'HKD':
                    converted_amount = float(value)*exchange_rate
                else:
                    sheet.cell(row=9+int(key[-2:])-len(delete_list), column=5).value = float(value)
            continue
        # Store converted amount in HKD
        if key in HKD_amount_header:
            if value and key[-2:] != delete_num:
                HKD_amount = value
            continue
        # Check HKD amount and evidence
        if key in evidence_header:
            if key[-2:] != delete_num:
                # Evidence exists
                if value:
                    # Currency is not HKD
                    if currency and currency != 'HKD':
                        evidence_index = str(int(key[-2:])-len(delete_list)).zfill(2)
                        evidence_dict[evidence_index] = value
                        if abs(float(HKD_amount) - converted_amount) > 250:
                            sheet.cell(row=9+int(key[-2:])-len(delete_list), column=5).value = converted_amount
                        else:
                            sheet.cell(row=9+int(key[-2:])-len(delete_list), column=5).value = float(HKD_amount)
                    # Currency has been changed to HKD
                    else:
                        # Delete converted HKD in Google Sheet
                        with open(os.path.join(config_path, 'converted_header.json')) as file:
                            converted_header_dict = json.load(file)
                        deleteCell(service, SPREADSHEET_ID, RANGE_NAME+'!'+converted_header_dict[key[-2:]]+str(index+2))
                        # Delete evidence both in Google Sheet and in Google Drive
                        try:
                            deleteDrive(config_path, value.split('=')[1])
                        except errors.HttpError as error:
                            print('An error occurred while deleting files in Google Drive: %s' % error)
                        with open(os.path.join(config_path, 'evidence_header.json')) as json_file:
                            evidence_header_dict = json.load(json_file)
                        deleteCell(service, SPREADSHEET_ID, RANGE_NAME+'!'+evidence_header_dict[key[-2:]]+str(index+2))
                # No evidence provided
                else:
                    if currency and currency != 'HKD':
                        sheet.cell(row=9+int(key[-2:])-len(delete_list), column=5).value = converted_amount
            # The item is marked as deleted
            else:
                # Evidence exists
                if value:
                    # Delete evidence both in Google Sheet and in Google Drive
                    try:
                        deleteDrive(config_path, value.split('=')[1])
                    except errors.HttpError as error:
                        print('An error occurred while deleting files in Google Drive: %s' % error)
                    with open(os.path.join(config_path, 'evidence_header.json')) as json_file:
                        evidence_header_dict = json.load(json_file)
                    deleteCell(service, SPREADSHEET_ID, RANGE_NAME+'!'+evidence_header_dict[key[-2:]]+str(index+2))
            continue
        # Write receipt URL
        if key in receipt_header:
            if key[-2:] != delete_num:
                receipt_index = str(int(key[-2:])-len(delete_list)).zfill(2)
                receipt_dict[receipt_index] = value
            # The item is marked as deleted
            else:
                # Delete receipt both in Google Sheet and in Google Drive
                try:
                    deleteDrive(config_path, value.split('=')[1])
                except errors.HttpError as error:
                    print('An error occurred while deleting files in Google Drive: %s' % error)
                with open(os.path.join(config_path, 'receipt_header.json')) as json_file:
                    receipt_header_dict = json.load(json_file)
                deleteCell(service, SPREADSHEET_ID, RANGE_NAME+'!'+receipt_header_dict[key[-2:]]+str(index+2))
            continue
    
    excel_name = 'Expense Claim Form_'+response_dict['Claim Number']+'.xlsx'
    excel_path = os.path.join(form_path, excel_name)
    pdf_name = 'Expense Claim Form_'+response_dict['Claim Number']+'.pdf'
    pdf_path = os.path.join(form_path, pdf_name)
    book.save(excel_path)
    excelToPDF(excel_path, pdf_path)
    return excel_name, excel_path, pdf_name, pdf_path, receipt_dict, evidence_dict, ver_num, has_row


# Rotate image
def rotateImage(image):
    try:
        '''
        for orientation in ExifTags.TAGS.keys():
            if ExifTags.TAGS[orientation] == 'Orientation':
                break
        exif = dict(image._getexif().items())

        if exif[orientation] == 3:
            image = image.rotate(180, expand=True)
        elif exif[orientation] == 6:
            image = image.rotate(270, expand=True)
        elif exif[orientation] == 8:
            image = image.rotate(90, expand=True)
        '''
        if image.height > image.width:
                image = image.rotate(270, expand=True)
        return image
    except (AttributeError, KeyError, IndexError):
        return


# Create batches for dict
def dictBatch(dict_items, batch_size):
    iterator = iter(dict_items)
    while True:
        result = tuple(itertools.islice(iterator, batch_size))
        if not result:
            break
        yield result


# Font Source: https://fonts.google.com/specimen/Roboto?selection.family=Roboto:wght@900#standard-styles
def imgToPDF(image_path, output_path, font_path, claim_num, ver_num, img_type, resample=Image.ANTIALIAS):
    # Collect images that match the claim number
    img_dict = dict((f.split('_')[2][:2], os.path.join(image_path, f)) for f in os.listdir(image_path) if os.path.isfile(os.path.join(image_path, f)) and f.split('_')[1] == claim_num and f.endswith('.jpg'))
    
    # Read valid images
    images = {}
    for item, img in img_dict.items():
        try:
            images[item] = Image.open(img)
        except IOError:
            continue
    
    if not images:
        print('No images currently.')
        return None, None
    
    # Rotate images if height > width
    for item, image in images.items():
        try:
            if image.height > image.width:
                images[item] = image.rotate(270, expand=True)
            else:
                continue
        except (AttributeError, KeyError, IndexError):
            continue
    
    # Generate combined images in batch
    batchsize = 4
    im_list_combined = []
    for batch in dictBatch(images.items(), batchsize):    
        # A4 size:
        # 72 ppi -> 595 x 842 pixel
        # 150 ppi -> 1240 x 1754 pixel
        # 600 ppi -> 4960 x 7016 pixel
        a4_width = 1240
        a4_height = 1754
        
        # Resize image according to A4 size
        width = a4_width - 150
        img_dict_resize = dict((image[0], image[1].resize((width, int(image[1].height*float(width)/image[1].width)), resample=resample)) for image in batch)
        total_height = sum(image.height for item, image in img_dict_resize.items())
        
        # Resize again if total height is larger than A4 height
        if total_height > a4_height:
            resize_portion = (a4_height-100)/total_height
            img_dict_resize = dict((item, image.resize((int(image.width*resize_portion), int(image.height*resize_portion)), resample=resample)) for item, image in img_dict_resize.items())
            total_height = sum(image.height for item, image in img_dict_resize.items())
        
        new_size = (a4_width, a4_height)
        new_color = (255, 255, 255, 0)
        new_image = Image.new('RGB', new_size, color=new_color)
        
        pos_y = 70
        for item, image in img_dict_resize.items():
            new_image.paste(image, (50, pos_y))
            
            draw = ImageDraw.Draw(new_image)
            
            # Draw Claim Number
            claim_number_font = ImageFont.truetype(os.path.join(font_path, 'Roboto', 'Roboto-LightItalic.ttf'), size=50)
            draw.multiline_text((50, 6), claim_num, font=claim_number_font, fill=(0, 0, 0))
            
            # Draw Version Number
            version_font = ImageFont.truetype(os.path.join(font_path, 'Roboto', 'Roboto-LightItalic.ttf'), size=50)
            draw.multiline_text((a4_width-200, 6), 'Ver.: '+str(ver_num), font=version_font, fill=(0, 0, 0))
            
            font = ImageFont.truetype(os.path.join(font_path, 'Roboto', 'Roboto-Bold.ttf'), size=50)
            # Starting position of the message
            size = (a4_width-70, pos_y)
            message = item
            color = 'rgb(0, 0, 0)' # Black color
    
            # Draw the message on the background
            draw.text(size, message, fill=color, font=font)
    
            pos_y += image.height
        im_list_combined.append(new_image)
    
    pdf_name = img_type+'_'+claim_num+'.pdf'
    pdf_path = os.path.join(output_path, pdf_name)
    if im_list_combined:
        # Save a single PDF with multiple pages
        im_list_combined[0].save(pdf_path, save_all=True, append_images=im_list_combined[1:])
    
    return pdf_name, pdf_path


def send_mail(config_path, applicant_email, applicant_name, supervisor_name, supervisors, filename_list='', filepath_list='', claim_num='', edit_url='', approve=False, correct=False, cancel=False, isTls=False):
    with open(os.path.join(config_path, 'email_info.json')) as file:
        email_dict = json.load(file)
        sender_email = email_dict['email']
        password = email_dict['password']
    main_form = <MAIN_GOOGLE_FORM_LINK>
    correction_form = <CORRECTION_GOOGLE_FORM_LINK>
    
    msg = MIMEMultipart()
    msg['Date'] = formatdate(localtime=True)
    msg['From'] = sender_email
    if cancel:
        receiver_mail = applicant_email
        msg['To'] = receiver_mail
        msg['Subject'] = 'Cancellation Notice - Expense Claim'
        text = 'Dear {},\n\n<CANCELLATION_MESSAGE_TO_APPLICANT>\nForm Link: {}\n\n<SIGN-OFF_AND_SIGNATURE_BLOCK>'.format(applicant_name, main_form)
    elif approve:
        receiver_mail = supervisors
        msg['To'] = ', '.join(receiver_mail)
        if correct:
            msg['Subject'] = 'Expense Claim Form Approval - Modified'
            text = 'Dear {},\n\n<MODIFICATION_MESSAGE_TO_SUPERVISOR>\n\n<SIGN-OFF_AND_SIGNATURE_BLOCK>'.format(supervisor_name, applicant_name)
        else:
            msg['Subject'] = 'Expense Claim Form Approval'
            text = 'Dear {},\n\n<REQUEST_APPROVAL_MESSAGE_TO_SUPERVISOR>\n\n<SIGN-OFF_AND_SIGNATURE_BLOCK>'.format(supervisor_name, applicant_name)
    elif correct:
        receiver_mail = applicant_email
        msg['To'] = receiver_mail
        msg['Subject'] = 'Modification Notice - Expense Claim'
        text = 'Dear {},\n\n<MODIFICATION_MESSAGE_TO_APPLICANT>\n\n<SIGN-OFF_AND_SIGNATURE_BLOCK>'.format(applicant_name)
    else:
        receiver_mail = applicant_email
        msg['To'] = receiver_mail
        msg['Subject'] = 'Success - Expense Claim'
        text = 'Dear {},\n\n<SUCCESS_MESSAGE_TO_APPLICANT>\n\n<SIGN-OFF_AND_SIGNATURE_BLOCK>'.format(applicant_name, claim_num, edit_url, correction_form)
    msg.attach(MIMEText(text))
    
    # Attach files
    for filename, filepath in zip(filename_list, filepath_list):
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(open(os.path.join(filepath), 'rb').read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename="{}"'.format(filename))
        msg.attach(part)
    
    # Log in to server using secure context and send email
    context = ssl.create_default_context()
    #with smtplib.SMTP_SSL(<SMTP_SERVER_NAME>, <SMTP_PORT_NUMBER>, context=context) as server:
    with smtplib.SMTP(<SMTP_SERVER_NAME>, <SMTP_PORT_NUMBER>) as server:
        if isTls:
            server.starttls(context=context)
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_mail, msg.as_string())
        server.quit()


def receiptEvidenceUpdate(config_path, service, response, index, sent):
    corrected = False
    try:
        # Get Google Spreadsheet data
        header_row_correct, response_list_correct = getSheetData(service, correct=True)
    except TypeError:
        print('The sheet for correction is empty.')
        return corrected, None
    
    if response_list_correct:
        for i, r in enumerate(response_list_correct):
            # Correction claim number matches the current claim number in response
            if r[3] == response[0]:
                # Check if the approval email for the response has been sent or not
                if sent == 'TRUE':
                    # Mark 'Finished' as True and skip it next time
                    updateCell(service, SPREADSHEET_CORRECT_ID, RANGE_NAME+'!A'+str(i+2), True)
                    continue
                else:
                    # The correction have not been done yet
                    if not r[0]:
                        response_dict_correct = dict(zip(header_row_correct, r))
                        
                        correct_receipt_header = ['Corrected Receipt-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
                        delete_evidence_header = ['Evidences of Exchange Rate to Delete [Evidence of Exchange Rate-{0}]'.format(str(i+1).zfill(2)) for i in range(13)]
                        correct_evidence_header = ['Corrected Evidence of Exchange Rate-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
                        with open(os.path.join(config_path, 'receipt_header.json')) as file:
                            receipt_header_dict = json.load(file)
                        with open(os.path.join(config_path, 'evidence_header.json')) as json_file:
                            evidence_header_dict = json.load(json_file)
                        
                        for k, v in response_dict_correct.items():
                            # Update receipt link both in response and in Google Sheet
                            if k in correct_receipt_header and v:
                                ID = list(range(16, 125, 9))[int(k[-2:])-1]  #int(k[-2:])+15+9*(int(k[-2:])-1)
                                # Remove original receipt from Google Drive
                                try:
                                    deleteDrive(config_path, response[ID].split('=')[1])
                                except errors.HttpError as error:
                                    print('An error occurred while deleting files in Google Drive: %s' % error)
                                # Replace original receipt URL with the new one
                                response[ID] = v
                                updateCell(service, SPREADSHEET_ID, RANGE_NAME+'!'+receipt_header_dict[k[-2:]]+str(index+2), v)
                            # Evidences to delete
                            if k in delete_evidence_header and v == 'Delete':
                                ID = list(range(18, 127, 9))[int(k[-3:-1])-1]
                                # Check if evidence exists
                                if response[ID]:
                                    # Remove original evidence from Google Drive
                                    try:
                                        deleteDrive(config_path, response[ID].split('=')[1])
                                    except errors.HttpError as error:
                                        print('An error occurred while deleting files in Google Drive: %s' % error)
                                    # Remove original evidence URl from Google Sheet
                                    deleteCell(service, SPREADSHEET_ID, RANGE_NAME+'!'+evidence_header_dict[k[-3:-1]]+str(index+2))
                                    # Clear up evidence URL in response
                                    response[ID] = ''
                            # Update evidence link both in response and in Google Sheet
                            if k in correct_evidence_header and v:
                                ID = list(range(18, 127, 9))[int(k[-2:])-1]  #int(k[-2:])+17+9*(int(k[-2:])-1)
                                # Check if evidence exists
                                if response[ID]:
                                    # Remove original evidence from Google Drive
                                    try:
                                        deleteDrive(config_path, response[ID].split('=')[1])
                                    except errors.HttpError as error:
                                        print('An error occurred while deleting files in Google Drive: %s' % error)
                                # Replace original evidence URL with the new one
                                response[ID] = v
                                updateCell(service, SPREADSHEET_ID, RANGE_NAME+'!'+evidence_header_dict[k[-2:]]+str(index+2), v)
                        
                        # Mark 'Finished' as True and skip it next time
                        updateCell(service, SPREADSHEET_CORRECT_ID, RANGE_NAME+'!A'+str(i+2), True)
                        corrected = True
            # Pass if claim numbers do not match
            else:
                continue
        return corrected, response
    else:
        print('There is no correction on receipt and evidence required.')
        return corrected, None


def main():
    # Set all relevant paths
    script_path = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_path, 'config')
    font_path = os.path.join(script_path, 'Fonts')
    template_path = os.path.join(script_path, 'Template')
    # Outputs
    output_path = os.path.join(script_path, 'output')
    createPath(output_path)
    # Form and Summary
    form_path = os.path.join(output_path, 'expense claim form')
    createPath(form_path)
    summary_path = os.path.join(output_path, 'summary')
    createPath(summary_path)
    # Receipt
    receipt_path = os.path.join(output_path, 'receipts')
    createPath(receipt_path)
    receipt_image_path = os.path.join(receipt_path, 'Image')
    createPath(receipt_image_path)
    receipt_output_path = os.path.join(receipt_path, 'PDF')
    createPath(receipt_output_path)
    # Evidence
    evidence_path = os.path.join(output_path, 'evidences')
    createPath(evidence_path)
    evidence_image_path = os.path.join(evidence_path, 'Image')
    createPath(evidence_image_path)
    evidence_output_path = os.path.join(evidence_path, 'PDF')
    createPath(evidence_output_path)
    
    # Get Google Sheets API credential
    creds = getCredential(config_path, 'sheet')
    # Call the Sheets API
    service = build('sheets', 'v4', credentials=creds)
    
    try:
    # Get Google Spreadsheet data
        header_row, response_list = getSheetData(service)
    except TypeError:
        print('The sheet is empty.')
        return
    
    # Check if any response exists
    if not response_list:
        print('There are no responses currently.')
        return
    
    # Read applicant's data from email list csv
    email_list = pd.read_csv(os.path.join(config_path, 'Email List.csv'))
    # Record deleted number of responses in Google Sheet for index
    deleted_num = 0
    
    # Create dictionary for each response
    for index, response in enumerate(response_list):
        submission_date = datetime.strptime(response[4], '%d/%m/%Y %H:%M:%S')
        applicant_email = response[5]
        applicant_name = email_list[email_list['Email'] == applicant_email]['First Name'].values[0]
        supervisor_name = email_list[email_list['Email'] == applicant_email]['Supervisor First Name'].values[0]
        supervisors = []
        supervisor_email = email_list[email_list['Email'] == applicant_email]['Supervisor Email'].values[0]
        supervisors.append(supervisor_email)
        
        # Delete response and all relevant outputs if cancelled
        if response[6] == 'Yes':
            # Send cancellation email to applicants
            send_mail(config_path, applicant_email, applicant_name, supervisor_name, supervisors, cancel=True)
            # Delete receipts from Google Drive
            for i in range(16, 125, 9):
                if response[i]:
                    try:
                        deleteDrive(config_path, response[i].split('=')[1])
                    except errors.HttpError as error:
                        print('An error occurred while deleting files in Google Drive: %s' % error)
            # Delete evidences from Google Drive
            for j in range(18, 127, 9):
                if response[j]:
                    try:
                        deleteDrive(config_path, response[j].split('=')[1])
                    except errors.HttpError as error:
                        print('An error occurred while deleting files in Google Drive: %s' % error)
            # Delete response from Google Sheet
            deleteRow(service, SPREADSHEET_ID, index+1-deleted_num)
            deleted_num += 1
            # Delete all relevant files in local file system
            deleteFile(form_path, response[0], 'Excel')
            deleteFile(form_path, response[0], 'PDF')
            deleteFile(receipt_image_path, response[0], 'Image')
            deleteFile(receipt_output_path, response[0], 'PDF')
            deleteFile(evidence_image_path, response[0], 'Image')
            deleteFile(evidence_output_path, response[0], 'PDF')
            continue
        
        # Check if claim number and claim date exist in Google Sheet
        # Write expense claim if both of them do not exist
        # 'Edited' column will be TRUE when new response occurs
        if not response[0] and not response[1]:
            # Generate claim number with timestamp, project code and applicant name
            submit_timestamp = int(time.mktime(submission_date.timetuple()))
            #claim_number = str(submit_timestamp) + '-' + str(uuid4())
            claim_number = applicant_name + '-' + response[9] + '-' + str(submit_timestamp)
            
            # Add claim number to the response and update Google Sheet
            response[0] = claim_number
            updateCell(service, SPREADSHEET_ID, RANGE_NAME+'!A'+str(index+2), claim_number)
            
            # Relate header to relevant value for each response
            response_dict = dict(zip(header_row, response))
            
            # Write individual claim form
            form_excel_name, form_excel_path, form_pdf_name, form_pdf_path, receipt_dict, evidence_dict, ver_num, has_row = writeClaimForm(service, index, config_path, template_path, form_path, response_dict)
            
            # Conduct the following steps if rows exist
            if has_row:
                # Download receipts
                downloadDrive(config_path, receipt_image_path, receipt_dict, response_dict['Claim Number'], 'Receipt')
                
                # Delete receipts with previous indices in local file system if needed
                deleteFile(receipt_image_path, response_dict['Claim Number'], 'Image', receipt_dict)
                
                filename_list = [form_excel_name, form_pdf_name]
                filepath_list = [form_excel_path, form_pdf_path]
                
                # Write receipts PDF
                receipt_pdf_name, receipt_pdf_path = imgToPDF(receipt_image_path, receipt_output_path, font_path, response_dict['Claim Number'], ver_num, 'Receipts')
                filename_list.append(receipt_pdf_name)
                filepath_list.append(receipt_pdf_path)
                
                if evidence_dict:
                    # Download evidences if exist
                    downloadDrive(config_path, evidence_image_path, evidence_dict, response_dict['Claim Number'], 'Evidence')
                    
                    # Delete evidences with previous indices in local file system if needed
                    deleteFile(evidence_image_path, response_dict['Claim Number'], 'Image', evidence_dict)
                    
                    # Write evidences PDF
                    evidence_pdf_name, evidence_pdf_path = imgToPDF(evidence_image_path, evidence_output_path, font_path, response_dict['Claim Number'], ver_num, 'Evidences')
                    filename_list.append(evidence_pdf_name)
                    filepath_list.append(evidence_pdf_path)
                else:
                    # Delete evidences downloaded in local file system previously if exist
                    deleteFile(evidence_image_path, response_dict['Claim Number'], 'Image')
                    deleteFile(evidence_output_path, response_dict['Claim Number'], 'PDF')
                
                # Send the claim form, receipts, evidences and claim number to applicant
                send_mail(config_path, applicant_email, applicant_name, supervisor_name, supervisors, filename_list=filename_list, filepath_list=filepath_list, claim_num=response_dict['Claim Number'], edit_url=response_dict['Form Response Edit URL'])
                # Update 'Edited' column to FALSE in Google Sheet
                updateCell(service, SPREADSHEET_ID, RANGE_NAME+'!C'+str(index+2), False)
            # The claim is considered as cancelled if no row exists
            else:
                # Send cancellation email to applicants
                send_mail(config_path, applicant_email, applicant_name, supervisor_name, supervisors, cancel=True)
                # Delete receipts and evidences from Google Drive
                for key, value in response_dict.items():
                    receipt_header = ['Upload Receipts-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
                    evidence_header = ['Evidence of Exchange Rate-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
                    if key in receipt_header:
                        if value:
                            try:
                                deleteDrive(config_path, value.split('=')[1])
                            except errors.HttpError as error:
                                print('An error occurred while deleting files in Google Drive: %s' % error)
                    if key in evidence_header:
                        if value:
                            try:
                                deleteDrive(config_path, value.split('=')[1])
                            except errors.HttpError as error:
                                print('An error occurred while deleting files in Google Drive: %s' % error)
                # Delete response from Google Sheet
                deleteRow(service, SPREADSHEET_ID, index+1-deleted_num)
                deleted_num += 1
                # Delete all relevant files in local file system
                deleteFile(form_path, response_dict['Claim Number'], 'Excel')
                deleteFile(form_path, response_dict['Claim Number'], 'PDF')
                deleteFile(receipt_image_path, response_dict['Claim Number'], 'Image')
                deleteFile(receipt_output_path, response_dict['Claim Number'], 'PDF')
                deleteFile(evidence_image_path, response_dict['Claim Number'], 'Image')
                deleteFile(evidence_output_path, response_dict['Claim Number'], 'PDF')
                continue
        # Only claim number exists
        elif response[0] and not response[1]:
            # The response is edited in the parts other than receipts and evidences
            if response[2] == 'TRUE':
                # Only allow modification within application day/before approval email is sent
                if response[3] == 'TRUE':
                    # Update 'Edited' column to FALSE in Google Sheet
                    updateCell(service, SPREADSHEET_ID, RANGE_NAME+'!C'+str(index+2), False)
                    continue
                else:
                    # Check if any correction on receipts and evidences required
                    corrected, response_new = receiptEvidenceUpdate(config_path, service, response, index, response[3])
                    
                    # Assign updated response if it is corrected
                    if corrected and response_new:
                        response = response_new
                    
                    # Write expense claim form, receipts and evidences again
                    # Relate header to relevant value for each response
                    response_dict = dict(zip(header_row, response))
                    
                    # Write individual claim form
                    form_excel_name, form_excel_path, form_pdf_name, form_pdf_path, receipt_dict, evidence_dict, ver_num, has_row = writeClaimForm(service, index, config_path, template_path, form_path, response_dict, update=True)
                    
                    # Conduct the following steps if rows exist
                    if has_row:
                        # Download receipts
                        downloadDrive(config_path, receipt_image_path, receipt_dict, response_dict['Claim Number'], 'Receipt')
                        
                        # Delete receipts with previous indices in local file system if needed
                        deleteFile(receipt_image_path, response_dict['Claim Number'], 'Image', receipt_dict)
                        
                        filename_list = [form_excel_name, form_pdf_name]
                        filepath_list = [form_excel_path, form_pdf_path]
                        
                        # Write receipts PDF
                        receipt_pdf_name, receipt_pdf_path = imgToPDF(receipt_image_path, receipt_output_path, font_path, response_dict['Claim Number'], ver_num, 'Receipts')
                        filename_list.append(receipt_pdf_name)
                        filepath_list.append(receipt_pdf_path)
                        
                        if evidence_dict:
                            # Download evidences if exist
                            downloadDrive(config_path, evidence_image_path, evidence_dict, response_dict['Claim Number'], 'Evidence')
                            
                            # Delete evidences with previous indices in local file system if needed
                            deleteFile(evidence_image_path, response_dict['Claim Number'], 'Image', evidence_dict)
                            
                            # Write evidences PDF
                            evidence_pdf_name, evidence_pdf_path = imgToPDF(evidence_image_path, evidence_output_path, font_path, response_dict['Claim Number'], ver_num, 'Evidences')
                            filename_list.append(evidence_pdf_name)
                            filepath_list.append(evidence_pdf_path)
                        else:
                            # Delete evidences downloaded in local file system previously if exist
                            deleteFile(evidence_image_path, response_dict['Claim Number'], 'Image')
                            deleteFile(evidence_output_path, response_dict['Claim Number'], 'PDF')
                    
                        # Send the claim form, receipts, evidences and claim number to applicant
                        send_mail(config_path, applicant_email, applicant_name, supervisor_name, supervisors, filename_list=filename_list, filepath_list=filepath_list, correct=True)
                        # Update 'Edited' column to FALSE in Google Sheet
                        updateCell(service, SPREADSHEET_ID, RANGE_NAME+'!C'+str(index+2), False)
                        # Update 'Sent' column to FALSE in Google Sheet
                        #updateCell(service, SPREADSHEET_ID, RANGE_NAME+'!D'+str(index+2), False)
                        # Update response in case applicants submit during sending approval email period
                        #response[3] = 'FALSE'
                    # The claim is considered as cancelled if no row exists
                    else:
                        # Send cancellation email to applicants
                        send_mail(config_path, applicant_email, applicant_name, supervisor_name, supervisors, cancel=True)
                        # Delete receipts and evidences from Google Drive
                        for key, value in response_dict.items():
                            receipt_header = ['Upload Receipts-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
                            evidence_header = ['Evidence of Exchange Rate-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
                            if key in receipt_header:
                                if value:
                                    try:
                                        deleteDrive(config_path, value.split('=')[1])
                                    except errors.HttpError as error:
                                        print('An error occurred while deleting files in Google Drive: %s' % error)
                            if key in evidence_header:
                                if value:
                                    try:
                                        deleteDrive(config_path, value.split('=')[1])
                                    except errors.HttpError as error:
                                        print('An error occurred while deleting files in Google Drive: %s' % error)
                        # Delete response from Google Sheet
                        deleteRow(service, SPREADSHEET_ID, index+1-deleted_num)
                        deleted_num += 1
                        # Delete all relevant files in local file system
                        deleteFile(form_path, response_dict['Claim Number'], 'Excel')
                        deleteFile(form_path, response_dict['Claim Number'], 'PDF')
                        deleteFile(receipt_image_path, response_dict['Claim Number'], 'Image')
                        deleteFile(receipt_output_path, response_dict['Claim Number'], 'PDF')
                        deleteFile(evidence_image_path, response_dict['Claim Number'], 'Image')
                        deleteFile(evidence_output_path, response_dict['Claim Number'], 'PDF')
                        continue
            # The response is not edited in main Google Sheet
            else:
                # Check if any correction on receipts and evidences required
                corrected, response_new = receiptEvidenceUpdate(config_path, service, response, index, response[3])
                
                # Assign updated response and write results again if it is corrected
                if corrected and response_new:
                    response = response_new
                
                    # Write expense claim form, receipts and evidences again
                    # Relate header to relevant value for each response
                    response_dict = dict(zip(header_row, response))
                    
                    # Write individual claim form
                    form_excel_name, form_excel_path, form_pdf_name, form_pdf_path, receipt_dict, evidence_dict, ver_num, has_row = writeClaimForm(service, index, config_path, template_path, form_path, response_dict, update=True)
                    
                    # Conduct the following steps if rows exist
                    if has_row:
                        # Download receipts
                        downloadDrive(config_path, receipt_image_path, receipt_dict, response_dict['Claim Number'], 'Receipt')
                        
                        # Delete receipts with previous indices in local file system if needed
                        deleteFile(receipt_image_path, response_dict['Claim Number'], 'Image', receipt_dict)
                        
                        filename_list = [form_excel_name, form_pdf_name]
                        filepath_list = [form_excel_path, form_pdf_path]
                        
                        # Write receipts PDF
                        receipt_pdf_name, receipt_pdf_path = imgToPDF(receipt_image_path, receipt_output_path, font_path, response_dict['Claim Number'], ver_num, 'Receipts')
                        filename_list.append(receipt_pdf_name)
                        filepath_list.append(receipt_pdf_path)
                        
                        if evidence_dict:
                            # Download evidences if exist
                            downloadDrive(config_path, evidence_image_path, evidence_dict, response_dict['Claim Number'], 'Evidence')
                            
                            # Delete evidences with previous indices in local file system if needed
                            deleteFile(evidence_image_path, response_dict['Claim Number'], 'Image', evidence_dict)
                            
                            # Write evidences PDF
                            evidence_pdf_name, evidence_pdf_path = imgToPDF(evidence_image_path, evidence_output_path, font_path, response_dict['Claim Number'], ver_num, 'Evidences')
                            filename_list.append(evidence_pdf_name)
                            filepath_list.append(evidence_pdf_path)
                        else:
                            # Delete evidences downloaded in local file system previously if exist
                            deleteFile(evidence_image_path, response_dict['Claim Number'], 'Image')
                            deleteFile(evidence_output_path, response_dict['Claim Number'], 'PDF')
                    
                        # Send the claim form, receipts, evidences and claim number to applicant
                        send_mail(config_path, applicant_email, applicant_name, supervisor_name, supervisors, filename_list=filename_list, filepath_list=filepath_list, correct=True)
                        # Update 'Edited' column to FALSE in Google Sheet
                        updateCell(service, SPREADSHEET_ID, RANGE_NAME+'!C'+str(index+2), False)
                        # Update 'Sent' column to FALSE in Google Sheet
                        #updateCell(service, SPREADSHEET_ID, RANGE_NAME+'!D'+str(index+2), False)
                        # Update response in case applicants submit during sending approval email period
                        #response[3] = 'FALSE'
                    # The claim is considered as cancelled if no row exists
                    else:
                        # Send cancellation email to applicants
                        send_mail(config_path, applicant_email, applicant_name, supervisor_name, supervisors, cancel=True)
                        # Delete receipts and evidences from Google Drive
                        for key, value in response_dict.items():
                            receipt_header = ['Upload Receipts-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
                            evidence_header = ['Evidence of Exchange Rate-{0}'.format(str(i+1).zfill(2)) for i in range(13)]
                            if key in receipt_header:
                                if value:
                                    try:
                                        deleteDrive(config_path, value.split('=')[1])
                                    except errors.HttpError as error:
                                        print('An error occurred while deleting files in Google Drive: %s' % error)
                            if key in evidence_header:
                                if value:
                                    try:
                                        deleteDrive(config_path, value.split('=')[1])
                                    except errors.HttpError as error:
                                        print('An error occurred while deleting files in Google Drive: %s' % error)
                        # Delete response from Google Sheet
                        deleteRow(service, SPREADSHEET_ID, index+1-deleted_num)
                        deleted_num += 1
                        # Delete all relevant files in local file system
                        deleteFile(form_path, response_dict['Claim Number'], 'Excel')
                        deleteFile(form_path, response_dict['Claim Number'], 'PDF')
                        deleteFile(receipt_image_path, response_dict['Claim Number'], 'Image')
                        deleteFile(receipt_output_path, response_dict['Claim Number'], 'PDF')
                        deleteFile(evidence_image_path, response_dict['Claim Number'], 'Image')
                        deleteFile(evidence_output_path, response_dict['Claim Number'], 'PDF')
                        continue
                # No any updates on the response
                else:
                    # Check if the approval email is sent or not
                    if response[3] == 'TRUE':
                        '''
                        # Check if any status updated in claim form
                        for filename in os.listdir(form_path):
                            if filename.endswith('.xlsx') and filename[:-5].split('_')[1] == response[0]:
                                book = openpyxl.load_workbook(os.path.join(form_path, filename), data_only=True)
                                sheet = book.active
                                status_list = []
                                for i in range(13):
                                    status_list.append(sheet.cell(row=10+i, column=6).value)
                                
                                # Skip the response if no status updated
                                if all(j is None for j in status_list):
                                    book.close()
                                    continue
                                # At least one status is updated
                                else:
                                    # check if status equals to 'Approved' or 'Rejected'
                                    if sum([True if status == 'Approved' or status == 'Rejected' else False for status in status_list]):
                                        
                                    # Skip the response if the status does not meet the format
                                    else:
                                        print('The status of Claim Number', response[0], 'does not meet the standard format.')
                                        book.close()
                                        continue
                        '''
                        # Check if claim date exists in claim form
                        for filename in os.listdir(form_path):
                            if filename.endswith('.xlsx') and filename[:-5].split('_')[1] == response[0]:
                                book = openpyxl.load_workbook(os.path.join(form_path, filename), data_only=True)
                                sheet = book.active
                                claim_date = sheet.cell(row=2, column=5).value
                        # If claim date exists in local Excel file
                        if claim_date:
                            # Convert datetime format to string
                            claim_date_str = claim_date.strftime('%d/%m/%Y')
                            # Update claim date both in Google Sheet and in response
                            response[1] = claim_date_str
                            updateCell(service, SPREADSHEET_ID, RANGE_NAME+'!B'+str(index+2), claim_date_str)
                            
                            # Do not write edit url in summaries
                            # Write overall summary
                            writeOverallSummary(summary_path, header_row[:-1], response[:-1])
                            
                            # Write response in month summary if it is in last month
                            #start_date = date.today().replace(day=1)  # First day of current month
                            #end_date = date.today().replace(day=calendar.monthrange(submission_date.year, submission_date.month)[1])  # Last day of current month
                            start_date = (date.today().replace(day=1) - timedelta(days=1)).replace(day=1)  # First day of last month
                            end_date = date.today().replace(day=1) - timedelta(days=1)  # Last day of last month
                            if claim_date.date() >= start_date and claim_date.date() <= end_date:
                                writeMonthSummary(summary_path, claim_date.year, claim_date.month, header_row[:-1], response[:-1])
                            
                            # Write response in audit summary if it is in last audit period
                            current_year = date.today().year
                            #next_year = date.today().year + 1
                            last_year = date.today().year - 1
                            audit_start_date = datetime(last_year, 4, 1)
                            audit_end_date = datetime(current_year, 3, 31)
                            if claim_date >= audit_start_date and claim_date <= audit_end_date:
                                writeAuditSummary(summary_path, last_year, current_year, header_row[:-1], response[:-1])
                            continue
                    # Record filenames and filepaths for approval email
                    else:
                        path_list = [form_path, receipt_output_path, evidence_output_path]
                        filename_list = []
                        filepath_list = []
                        # Collect Excel filename and filepath
                        for filename in os.listdir(form_path):
                            if filename.endswith('.xlsx') and filename[:-5].split('_')[1] == response[0]:
                                filename_list.append(filename)
                                filepath_list.append(os.path.join(form_path, filename))
                        # Collect PDF filename and filepath
                        for path in path_list:
                            for filename in os.listdir(path):
                                if filename.endswith('.pdf') and filename[:-4].split('_')[1] == response[0]:
                                    filename_list.append(filename)
                                    filepath_list.append(os.path.join(path, filename))
                    continue
        # It is not possible to have claim date without claim number,
        # since expense claim form needs to be written and approved first.
        elif not response[0] and response[1]:
            continue
        # Do nothing if both claim number and claim date exist
        elif response[0] and response[1]:
            continue
        
        # Send the claim form, receipts, evidences and claim number to supervisors during specific period
        # The script is set to run from 00:00 every 15 mins
        # The period for sending approval email is from 00:00 to 01:00
        startTime = datetime.strptime('2020/1/1 00:00:00', '%Y/%m/%d %H:%M:%S').time()
        endTime = datetime.strptime('2020/1/1 01:00:00', '%Y/%m/%d %H:%M:%S').time()
        if startTime <= datetime.now().time() <= endTime:
            # New response has not been sent yet
            if not response[3]:
                send_mail(config_path, applicant_email, applicant_name, supervisor_name, supervisors, filename_list=filename_list, filepath_list=filepath_list, approve=True)
                # Update 'Sent' column to TRUE in Google Sheet
                updateCell(service, SPREADSHEET_ID, RANGE_NAME+'!D'+str(index+2), True)
            # Response is edited and has not been sent yet
            elif response[3] == 'FALSE':
                send_mail(config_path, applicant_email, applicant_name, supervisor_name, supervisors, filename_list=filename_list, filepath_list=filepath_list, approve=True, correct=True)
                # Update 'Sent' column to TRUE in Google Sheet
                updateCell(service, SPREADSHEET_ID, RANGE_NAME+'!D'+str(index+2), True)


if __name__ == '__main__':
    main()
